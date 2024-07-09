

from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'aula334_workbook.xlsx'

workbook = Workbook()
worksheet: Worksheet = workbook.active

# Criando os cabeçalhos
worksheet.cell(1,1,'Nome')
worksheet.cell(1,2,'Idade')
worksheet.cell(1,3,'Nota')

students = [
    #nome       idade   nota
    ['João',    14,     5.5],
    ['Maria',   15,     6.5],
    ['Paulo',   16,     7.5],
    ['Cris',    14,     8.5],
    ['Julio',   17,     9.5]
]

# for i in range(2, 10):
#     for j in range(1,4):
#         print(f'Linha {i} Coluna {j}' )

# LOGICA 1
# for i, student_row in enumerate(students, start=2):
#     # print(i, student_row)
#     for j, student_col in enumerate(student_row, start=1):
#         # print(i, j, student_col)
#         worksheet.cell(i, j, student_col)

#LOGICA 2
for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)